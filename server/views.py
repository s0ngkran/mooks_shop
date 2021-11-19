from django.http.response import HttpResponse
from datetime import date, timedelta
import openpyxl
from django.shortcuts import render, reverse
import datetime
from django.utils import timezone
from .utils import *
from .forms import *
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.views.generic.edit import CreateView
import pytz
from django.utils.timezone import make_aware


class RegisterPage(MyView):
    template_name = 'server/register.html'
    form_user = UserForm
    form_profile = ProfileForm

    def get(self, request, *args, **kwargs):
        if request.user.is_authenticated:
            return redirect('index-page')

        self.context.update({
            'form_user': self.form_user,
            'form_profile': self.form_profile,
        })

        return self.render(request)

    def post(self, request, *args, **kwargs):
        data = request.POST.copy()
        act = data.get('act')
        if act == 'register':
            form1 = self.form_user(data)
            form2 = self.form_profile(data)
            if form1.is_valid and form2.is_valid:
                if User.objects.filter(username=data['username']).exists():
                    messages.warning(request, 'duplicate username')
                    return redirect('register-page')
                new_user = User.objects.create_user(
                    username=data['username'],
                    password=data['password'],
                )
                new_user.save()
                new_profile = form2.save(commit=False)
                new_profile.user = new_user
                new_profile.save()

        user = authenticate(
            username=data['username'], password=data['password'])
        if user is not None:
            login(request, user)
            return redirect('index-page')

        return redirect('login-page')


class LoginPage(MyView):
    template_name = 'server/login.html'

    def get(self, request, *args, **kwargs):
        # Profile.objects.all().delete()
        # self.context.update({
        #     'users': User.objects.all(),
        #     'profiles': Profile.objects.all(),
        # })
        # messages.warning(request, 'test message')
        if request.user.is_authenticated:
            return redirect('index-page')
        return self.render(request)

    def post(self, request, *args, **kwargs):
        data = request.POST.copy()
        print('dat', data)
        user = authenticate(
            username=data['username'], password=data['password'])
        if user is not None:
            # do login
            login(request, user)

            # create profile if have no
            if user.profile_set.all().count() == 0:
                # create new profile
                Profile.objects.create(user=user)
                return redirect('index-page')
            elif user.profile_set.all().count() == 1:
                return redirect('index-page')
        messages.warning(request, 'not correct!')
        return redirect('login-page')


class LogoutPage(MyView):
    def get(self, request, *args, **kwargs):
        logout(request)
        return redirect('login-page')


class IndexPage(MyView):
    template_name = 'server/index.html'
    permission = 9999

    @has_perm
    def get(self, request, *args, **kwargs):
        return self.render(request)


class AddNewProductPage(MyView):
    template_name = 'server/add_product.html'
    permission = 9999

    @has_perm
    def get(self, request, *args, **kwargs):
        categories = Category.objects.all()

        def get_current_category():
            current_category = request.user.profile_set.all().first().current_category
            if current_category == None or current_category == '':
                current_category = 'Select Category...'
            return current_category

        self.context.update({
            'current_category':  get_current_category(),
            'categories': categories,
        })
        return self.render(request)

    @has_perm
    def post(self, request, *args, **kwargs):
        data = request.POST

        class DupError(Exception):
            def __str__(self):
                return """Duplicate Barcode"""
        # create
        new_product = None
        e = None
        try:
            def get_price(price):
                try:
                    price = float(price)
                    return price
                except:
                    return None

            def get_inventory(num):
                try:
                    num = int(num)
                    return num
                except:
                    return 0

            def get_category(id):
                ans = Category.objects.filter(id=id)
                if len(ans) == 1:
                    return id
                else:
                    return None

            def get_barcode(barcode):
                product = Product.objects.filter(barcode=barcode)
                if len(product) == 0:
                    return barcode
                else:
                    raise DupError()

            category = get_category(data.get('category'))
            price = get_price(data.get('price'))
            inventory = get_inventory(data.get('inventory'))
            barcode = get_barcode(data.get('barcode'))

            new_product = Product.objects.create(
                category_obj_id=category,
                barcode=barcode,
                description=data.get('description'),
                name=data.get('name'),
                price=price,
                inventory=inventory,
            )
            messages.success(request, new_product.barcode +
                             "-"+new_product.name)

        except DupError as e:
            self.handle_error(request, e)
        except Exception as e:
            self.handle_error(request, e)

        if new_product != None:
            profile = request.user.profile_set.all().first()
            profile.current_category = new_product.category_obj
            profile.save()
            self.context = {}
        return self.get(request)

    def handle_error(self, request, e):
        if e != None:
            self.context.update({
                'error': str(e)
            })
            return self.render(request)


class CashierPage(MyView):
    template_name = 'server/cashier.html'
    permission = 9999

    def get_total(self, transaction):
        # calculate price
        subs = SubTransaction.objects.filter(transaction_obj=transaction)
        total_promotion = 0
        total_item = 0
        total_pure = 0
        #### group promotion (kill group first)
        remaining_subs_from_group, group_price, discount_on_group = self.get_discount_from_promotion_on_group(subs)
        for sub in remaining_subs_from_group:
            total_pure += sub.n_item * sub.product_obj.price
            ##### calculate single promotion
            if sub.is_adjust == True:
                sub.price = sub.price
            else:
                sub.price = self.get_total_price_from_promotion(
                    sub.product_obj, sub.n_item, sub.product_obj.price)
            sub.save()
            total_promotion += sub.price
            total_item += sub.n_item

        # sumarize
        total_item = int(total_item)
        total_promotion = total_promotion + group_price

        transaction.discount_from_promotion = total_pure - total_promotion # not include promotion-group
        transaction.discount_from_promotion_on_group = discount_on_group
        transaction.total_item = total_item
        # not include is_adjust case
        if transaction.is_adjust == False:
            transaction.total = total_promotion
            transaction.save()
        return subs, total_promotion, total_item, transaction

    def get_discount_from_promotion_on_group(self, subs):
        # get all_products with promotion-group
        subs_with_promotion_list = []
        all_pros = []
        all_pro_ids = []
        remaining_subs = []
        total_pure = 0
        for sub in subs:
            pros = PromotionOnGroup.objects.filter(products=sub.product_obj.id)
            if pros.exists():
                assert len(pros) == 1
                pro = pros[0]
                total_pure += sub.n_item * sub.product_obj.price
                subs_with_promotion_list.append(sub)
                # avoid dup pro in all_pros
                if pro.id not in all_pro_ids:
                    all_pros.append(pro)
                    all_pro_ids.append(pro.id)
            else:
                remaining_subs.append(sub)
        '''
        _ = [
             (pro1, [sub1, sub2] )
             (pro1, [sub1, sub2] )
             (pro1, [sub1, sub2] )
        ]
        '''
        cat = []
        for pro in all_pros:
            sub_list = []
            # find sub in pro or not
            for sub in subs_with_promotion_list:
                sub.price = 0
                sub.save()
                if sub.product_obj in pro.products.all():
                    sub_list.append(sub)
            cat.append((pro, sub_list))
        
        #################################
        #################################
        #################################
        # this code treat that the prices in a group are the same
        # if not you need to sort by their price 
        #################################
        #################################
        #################################

        # cal group price and discount
        group_price = 0
        for pro, sub_list in cat:
            # sum n_item
            n_item = 0
            for sub in sub_list:
                n_item += sub.n_item
            
            # loop from higher number of n_item
            for pricing in pro.pricings.all().order_by('-price'):
                if n_item >= pricing.n_item and n_item > 0:
                    # get group price of each step
                    group_price += int(n_item/pricing.n_item) * pricing.price
                    n_item -= int(n_item/pricing.n_item) * pricing.n_item
            
            # cal remaining n_item
            group_price += n_item * sub.product_obj.price # last price from sub  ###### need to solve

        discount = total_pure - group_price
        return remaining_subs, group_price, discount
        

    def get_total_price_from_promotion(self, product_obj, n_item, price):
        # calculate single promotion
        all_promotions = Promotion.objects.filter(
            product_obj=product_obj).order_by('-n_item')
        if len(all_promotions) >= 1:
            total_price = 0
            for promotion in all_promotions:
                n_item_ = promotion.n_item
                price_ = promotion.price
                for i in range(100):
                    if n_item >= n_item_:
                        # kill from last
                        n_item -= n_item_
                        total_price += price_
                    else:
                        break
            # remains
            total_price += price * n_item

        elif len(all_promotions) == 0:
            total_price = price * n_item
        else:
            1/0
        return total_price

    def get_transaction(self, profile):
        # check transaction
        transaction = profile.current_transaction
        if transaction == None:
            # create transaction
            transaction = Transaction()
            transaction.save()
            profile.current_transaction = transaction
            profile.save()
        return transaction

    @has_perm
    def get(self, request, *args, **kwargs):
        profile = request.user.profile_set.all().first()
        if profile == None:
            return redirect('logout-page')
        transaction = self.get_transaction(profile)

        # subs, total, total_item, transaction = self.get_total(transaction)

        # update status
        mode = ''
        if transaction.status == None:
            mode += 'get money btn'
        self.context = {}
        self.context.update({
            'mode': mode,
        })
        return self.render(request)

    def render(self, request):
        # get transaction
        profile = request.user.profile_set.all().first()
        transaction = self.get_transaction(profile)

        # append phase
        if transaction.received == None:
            phase = 'start'
        else:
            phase = 'wait fin'

        # get sub
        subs, total, total_item, transaction = self.get_total(transaction)

        # append context
        self.context.update({
            'phase': phase,
            'transaction': transaction,
            'subs': subs.order_by('-updated_on'),
            'len_subs': len(subs),
        })
        return super().render(request)

    @has_perm
    def post(self, request, *args, **kwargs):
        # get data
        data = request.POST
        act = data.get('act')

        # get current transaction
        profile = request.user.profile_set.all().first()
        transaction = self.get_transaction(profile)


        if act == 'scan barcode':
            # find product
            barcode = request.POST.get('barcode')
            if barcode == '' or barcode == None:
                self.context.update({
                    'error': 'something wrongxx'
                })
                assert False

            # check n product
            products = Product.objects.filter(barcode=barcode)

            if len(products) == 0:
                self.context.update({
                    'error': 'bad barcode <br> <h1>Get Money</h1>'
                })
                return self.render(request)
            if len(products) != 1:
                self.context.update({
                    'error': 'duplicate product code<br>edit your product'
                })
                return self.render(request)
            product = products.first()

            #  is bank?
            if product.is_bank == True:
                transaction.bank = product
                transaction.save()
                return self.render(request)

            # check same product appear in transaction
            subs = SubTransaction.objects.filter(transaction_obj=transaction)
            same_subs = subs.filter(product_obj=product)
            if len(same_subs) >= 2:
                # something wrong
                # sub should has only one obj
                self.context.update({
                    'error': 'something wrong'
                })
                assert False

            elif len(same_subs) == 1:
                # duplicate product
                sub = same_subs.first()
                sub.n_item += 1
                sub.save()
            else:
                # new sub
                # create subtransaction
                sub = SubTransaction.objects.create(
                    transaction_obj=transaction,
                    product_obj=product,
                    n_item=1,
                    # created_on = timezone.now(),
                )
            self.context = {}
            return self.get(request)

        if act == 'adjust total':
            try:
                adj_price = float(data.get('price'))
            except:
                return self.get(request)

            transaction.total = adj_price
            transaction.is_adjust = True
            transaction.save()
            return self.get(request)

        if act == 'adjust sub':
            sub_id = data.get('sub_id')
            price = data.get('price')

            # clean data
            if sub_id == '' or sub_id == None:
                return self.get(request)
            try:
                price = float(price)
            except:
                return self.get(request)

            # adjust sub
            subs = SubTransaction.objects.filter(id=sub_id)
            if len(subs) == 1:
                sub = subs.first()
                sub.is_adjust = True
                sub.price = price
                sub.save()
            else:
                assert False

            return self.get(request)
        if act == 'get money':
            cash_amount = data.get('amount')
            online_amount = data.get('online-amount')

            def get_amount(num):
                try:
                    num = float(num)
                    return num
                except:
                    return 0
            cash_amount = get_amount(cash_amount)
            online_amount = get_amount(online_amount)
            cash_and_online_amount = cash_amount + online_amount

            # cal total
            total = transaction.total

            # check amount
            if cash_and_online_amount < total:
                self.context.update({
                    'error': 'not enough money',
                })
                return self.render(request)

            # update received
            transaction.received = cash_and_online_amount
            transaction.received_cash = cash_amount
            transaction.received_online = online_amount
            transaction.balance = cash_and_online_amount - total
            transaction.save()

            subs = transaction.subtransaction_set.all()
            mode = ''
            mode += 'received'
            self.context.update({
                'mode': mode,
                'transaction': transaction,
                'subs': subs,
                'len_subs': len(subs),
            })
            return self.render(request)

        if act == 'save':
            transaction.is_success = True
            transaction.save()
            new = Transaction.objects.create()
            profile.current_transaction = new
            profile.save()
            self.context = {}
            return redirect('cashier-page')
        
        if act == 'del item':
            subs = SubTransaction.objects.filter(pk=data.get('pk')).delete()

            return self.render(request)
        if act == 'clear all':
            transaction.delete()
            new = Transaction.objects.create()
            profile.current_transaction = new
            profile.save()
            return self.render(request)
        if act == 'edit number':
            num = data.get('num')
            try:
                num = int(num)
            except:
                self.context.update({
                    'error': 'not correct number',
                })
                return self.render(request)
            subs = SubTransaction.objects.filter(pk=data.get('pk'))
            if len(subs) != 1:
                self.context.update({
                    'error': 'something went wrong',
                })
                return self.render(request)
            sub = subs.first()
            sub.n_item = num
            sub.save()
            return self.render(request)



class CategoryCreateView(CreateView):
    model = Category
    fields = ['name', 'description']
    success_url = '/add-new-product/'


class PromotionCreateView(MyView):
    template_name = 'server/promotion_form.html'
    permission = 9999

    @has_perm
    def get(self, request, *args, **kwargs):
        return self.render(request)

    @has_perm
    def post(self, request, *args, **kwargs):
        data = request.POST
        barcode = data.get('barcode')
        try:
            n_item = int(data.get('n_item'))
            price = float(data.get('price'))
        except:
            return redirect('add-new-promotion-page')

        # change barcode to product
        products = Product.objects.filter(barcode=barcode)
        if len(products) != 1:
            # messages.warning(request, 'have no product')
            return redirect('add-new-promotion-page')
        product = products.first()

        # create new promotion
        _ = {
            'product_obj': product,
            'n_item': n_item,
            'price': price,
        }
        new_promotion = Promotion.objects.create(
            **_
        )
        print('new promotion is created')
        return self.render(request)


class TransactionPage(MyView):
    template_name = 'server/transaction.html'
    permission = 9999

    def get_time_diff(self, day_from_now=0):
        now = datetime.datetime.utcnow().replace(tzinfo=timezone.utc)
        timediff = now - day_from_now
        return timediff

    def get_aware_date(self,date_string):
        date_format = '%Y-%m-%d'
        unaware_date = datetime.datetime.strptime(date_string, date_format)
        aware_date = pytz.utc.localize(unaware_date)
        return aware_date
    def get_naive_date(self,date_string):
        date_format = '%Y-%m-%d'
        unaware_date = datetime.datetime.strptime(date_string, date_format)
        # aware_date = pytz.utc.localize(unaware_date)
        return unaware_date

    @has_perm
    def get(self, request, *args, **kwargs):
        today = timezone.now()
        # today = make_aware(today)
        yesterday = today - timezone.timedelta(days=1)
        yesterday = yesterday.date()
        date = kwargs.get('from_post')

        if date:
            date = self.get_naive_date(date)
            transactions = Transaction.objects.filter(
                updated_on__gte=date,  updated_on__lte=date+timezone.timedelta(days=1), is_success=True).order_by('-updated_on')
        else:
            transactions = Transaction.objects.filter(
                updated_on__gte=yesterday,  is_success=True).order_by('-updated_on')
        sum_total = 0
        sum_balance = 0
        sum_received = 0
        sum_received_cash = 0
        sum_received_online = 0
        for i, transaction in enumerate(transactions):
            transaction.i = i+1
            try:
                sum_total += transaction.total
                sum_balance += transaction.balance
                sum_received += transaction.received
                sum_received_cash += transaction.received_cash
                sum_received_online += transaction.received_online
            except:
                pass
        self.context.update({
            'transactions': transactions,
            'n_transaction': len(transactions),
            'sum_total': sum_total,
            'sum_balance': sum_balance,
            'sum_received': sum_received,
            'sum_received_cash': sum_received_cash,
            'sum_received_online': sum_received_online,
        })
        return self.render(request)

    @has_perm
    def post(self, request, *args, **kwargs):
        data = request.POST
        date = data.get('date')
        _=    {
                'from_post': date
            }
        return self.get(request, **_)

class SubTransactionPage(MyView):
    template_name = 'server/subtransaction.html'
    permission = 9999

    @has_perm
    def get(self, request, id, *args, **kwargs):
        transactions = Transaction.objects.filter(id=id)
        if len(transactions) != 1:
            return self.render(request)

        transaction = transactions.first()
        subs = transaction.subtransaction_set.all()
        self.context.update({
            'transaction': transaction,
            'subs': subs,
        })
        return self.render(request)


class MigratePage(MyView):
    # template_name = 'server/subtransaction.html'
    permission = 9999
    def migrate2(self, xlsx_path):
        wb_obj = openpyxl.load_workbook(xlsx_path)
        count = 0
        for sheet in wb_obj.worksheets:
            # row and column index start at 1 not 0
            for i in range(2, 200000):
                try:
                    category  = sheet.cell(row=i, column=1).value
                    barcode  = sheet.cell(row=i, column=2).value
                    # barcode = str(barcode) # add this line
                    # barcode = float(barcode.split('e+')[0]) * 10 ** int(barcode.split('e+')[1])
                    if 'e+' in str(barcode):
                        barcode = '%d'%(int(barcode))
                    else:
                        barcode = str(barcode)
                    quantity  = sheet.cell(row=i, column=3).value
                    comment  = sheet.cell(row=i, column=4).value
                    product_name  = sheet.cell(row=i, column=5).value
                    price  = sheet.cell(row=i, column=6).value
                    if category == None and price == None:
                        break
                    barcode = barcode.replace(' ', '')
                    try:
                        quantity = int(quantity)
                    except:
                        quantity = 0
                    try:
                        price = float(price)
                    except:
                        price = 0
                    # check cat
                    cats = Category.objects.filter(name=category)
                    if len(cats) == 0:
                        cat = Category.objects.create(
                            name=category
                        )
                    elif len(cats) == 1:
                        cat = cats.first()

                    # check has product?
                    products = Product.objects.filter(barcode=barcode)
                    if len(products) != 0:
                        continue

                    # create
                    product = Product.objects.create(
                        category_obj=cat,
                        barcode=barcode,
                        name=product_name,
                        price=price,
                        inventory=quantity,
                        description = comment,
                    )
                    # print('added', i)
                    count += 1
                except Exception as e:
                    print(i, e)
                    pass
        return count

        

    def migrate(self, data):
        count = 0
        for i, dat in enumerate(data.split('\n')):
            dat = dat.split('\t')
            try:
                category = dat[0]
                barcode = dat[1]
                quantity = dat[2]
                comment = dat[3]
                product_name = dat[4]
                price = dat[5]
                barcode = barcode.replace(' ', '')
                try:
                    quantity = int(quantity)
                except:
                    quantity = 0
                try:
                    price = float(price)
                except:
                    price = 0
                # check cat
                cats = Category.objects.filter(name=category)
                if len(cats) == 0:
                    cat = Category.objects.create(
                        name=category
                    )
                elif len(cats) == 1:
                    cat = cats.first()

                # check has product?
                products = Product.objects.filter(barcode=barcode)
                if len(products) != 0:
                    continue

                # create
                product = Product.objects.create(
                    category_obj=cat,
                    barcode=barcode,
                    name=product_name,
                    price=price,
                    inventory=quantity,
                    description = comment,
                )
                # print('added', i)
                count += 1

            except Exception as e:
                # print('error', e)
                pass
        return count

    @has_perm
    def get(self, request, *args, **kwargs):
        # from .m import get_data
        # count = self.migrate(get_data())
        xlsx_path = './data.xlsx'
        count  = self.migrate2(xlsx_path)
        return HttpResponse('success'+str(count))

class DoUpdate(MyView):
    template_name = 'server/update.html'
    permission = 9999
    @has_perm
    def post(self, request, *args, **kwargs):
        act = request.POST.get('act')
        if act == 'update':
            import os
            from shutil import copyfile
            # os.popen('git stash --include-untracked').read()
            # os.popen('git reset --hard').read()
            # os.popen('git clean -fd').read()
            os.popen('mkdir tempxx').read()
            os.popen('git clone https://github.com/s0ngkran/mooks_shop tempxx/mooks_shop').read()
            ######################
            # do update here
            for _,__, fnames in os.walk('tempxx/mooks_shop/server/templates/server/'):
                print('walk template/server/')
            for fname in fnames:
                # os.popen('cp tempxx/mooks_shop/server/templates/server/%s server/templates/server/%s'%(fname, fname)).read()
                copyfile('tempxx/mooks_shop/server/templates/server/%s'%fname, 'server/templates/server/%s'%fname)

            # replace react
            copyfile('tempxx/mooks_shop/frontend/static/frontend/main.js', './frontend/static/frontend/main.js')
            print('success main.js -.-')

            # replace urls.py
            for name in ['urls.py', 'utils.py', 'sers.py', 'models.py','apis.py', 'admin.py', 'views.py']:
                copyfile('tempxx/mooks_shop/server/%s'%name, 'server/%s'%name)

            # delete
            unique_name = str(datetime.datetime.now()).replace(' ','').replace(':', '').replace('.','').replace('-','')
            os.rename('tempxx/mooks_shop', 'tempxx/%s'%unique_name)

            #############################
            features =[
                'do this eiei oo',
                'by thieees',
            ]
            features = '<br>'+ '<br>'.join(features)
            self.context.update(
                {
                    'text': 'now you are in v2.4' + features,
                    'message': 'success',
                    'is_show_update_button': False,
                }
            )
            return self.render(request)
    @has_perm
    def get(self, request, *args, **kwargs):
        self.context.update(
            {
                'text': 'now you are in v2.4',
                'message': '',
                'is_show_update_button': True,
            }
        )
        return self.render(request)

class OnlineAmount(MyView):
    template_name = 'server/online_amount.html'
    permission = 9999

    @has_perm
    def get(self, request, *args, **kwargs):
        bank = kwargs.get('bank')
        
        banks = Product.objects.filter(is_bank=True)
        if kwargs.get('bank') == None:
            try:
                bank = banks.first()
            except:pass

        
        all_trans = Transaction.objects.filter(bank=bank)
        call_back = 5
        enddate = date.today() - timedelta(days=call_back)
        x_data = []
        y_data = []
        for i in range(call_back+1):
            startdate = enddate
            enddate = startdate + timedelta(days = 1)
            trans = all_trans.filter(updated_on__range = [startdate, enddate])
            value = 0
            for tran in trans:
                if tran.received_online != None:
                    value += tran.received_online
            y_data.append(float(value))
            x_data.append(startdate.strftime('%Y-%m-%d'))

        bank.x_data = x_data
        bank.y_data =y_data
        self.context.update({
            'banks': banks,
            'bank': bank,
        })
        return self.render(request)

    def find_transaction(self, request):
        profile = request.user.profile_set.all().first()
        if profile == None:
            return redirect('logout-page')
        transaction = CashierPage().get_transaction(profile)
        return profile, transaction

    @has_perm
    def post(self, request, *args, **kwargs):
        data = request.POST
        act = data.get('act')
        if act == 'search':
            bank = Product.objects.get(is_bank=True, pk=data.get('pk'))
            print('bakn', bank)
            return self.get(request, **{'bank': bank})